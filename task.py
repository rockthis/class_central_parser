from openpyxl import Workbook
from selenium import webdriver
from time import sleep

# driver = webdriver.Chrome('/Users/kostyafrolov/Downloads/chromedriver')   mac

driver = webdriver.Chrome('C:\\Users\\Admin\\Downloads\\chromedriver')    #windows

driver.get('https://www.class-central.com/subject/data-science')

sleep(4)

while True:                                                             # get all courses
    if driver.find_element_by_xpath('//*[@id="show-more-courses"]').is_displayed():
        driver.find_element_by_xpath('//*[@id="show-more-courses"]').click()
        sleep(6)
        if driver.find_element_by_xpath('//*[@id="signupModal-ask_for_signup"]/div/div/a').is_displayed():
            driver.find_element_by_xpath('//*[@id="signupModal-ask_for_signup"]/div/div/a').click()
            sleep(3)
    else:
        break


course_rate = []
course_name = []
course_provider = []
course_platform = []
course_start_date = []
course_url = []

course_rates = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]/td[4]')

course_names = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]/td[2]/a/span')

course_providers = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]')

course_platforms = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]/td[2]/ul/a')

course_start_dates = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]/td[3]')

course_urls = driver.find_elements_by_xpath('//*[@itemtype="http://schema.org/Event"]/td[2]/a')

for el in course_urls:
    course_url.append(el.get_attribute('href'))

for el in course_start_dates:
    course_start_date.append(el.get_attribute('textContent'))

for el in course_platforms:
    course_platform.append(el.get_attribute('textContent'))

for el in course_providers:
     providers = el.find_elements_by_xpath('td[2]/ul/li[1]/a')
     if providers:
        for temp in providers:
            course_provider.append(temp.get_attribute('textContent'))
     else:
         course_provider.append('No provider')
     scndprovider = el.find_elements_by_xpath('td[2]/ul/li[2]/a')
     if scndprovider:
         for temp in scndprovider:
             course_provider[-1] = course_provider[-1] + ', ' + temp.get_attribute('textContent')

for el in course_rates:
    course_rate.append(el.get_attribute('data-timestamp')[0:3])

for el in course_names:
    course_name.append(el.get_attribute('textContent'))

wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Course Name'
sheet.cell(row=1, column=2).value = 'Provider'
sheet.cell(row=1, column=3).value = 'Platform'
sheet.cell(row=1, column=4).value = 'Start Data'
sheet.cell(row=1, column=5).value = 'Rating'
sheet.cell(row=1, column=6).value = 'Course URL'
for i in range(0,len(course_name)):
    sheet.cell(row=2 + i, column=1).value = course_name[i]
for i in range(0,len(course_provider)):
    sheet.cell(row=2 + i, column=2).value = course_provider[i]
for i in range(0,len(course_rate)):
    sheet.cell(row=2 + i, column=5).value = course_rate[i]
for i in range(0,len(course_platform)):
    sheet.cell(row=2 + i, column=3).value = course_platform[i]
for i in range(0,len(course_start_date)):
    sheet.cell(row=2 + i, column=4).value = course_start_date[i]
for i in range(0,len(course_url)):
    sheet.cell(row=2 + i, column=6).value = course_url[i]
wb.save('class_central.xlsx')


driver.quit()


